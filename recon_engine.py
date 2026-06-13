"""
recon_engine.py — Core Reconciliation Engine for ReconAgent
Implements all 6 steps from the process draft document.
"""

import pandas as pd
import numpy as np
import json
import os
import io
from pathlib import Path
from datetime import datetime, timedelta, time
from typing import Optional, Dict, Tuple, List
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# PARTNER CONFIG MATRIX
# ─────────────────────────────────────────────────────────────────────────────

PARTNER_CONFIG: Dict[str, dict] = {
    "Apple": {
        "file_identifier": "zalopay_amp",          # lowercase for matching
        "cutoff_time": None,                        # file-based, no fixed cutoff
        "buffer_window_minutes": 0,
        "file_delivery_type": "DAILY_SPLIT",
        "tolerance_vnd": 0,
        "cancel_key_field": "merchant_txn_id",     # lowercase normalized
        "amount_field_hints": ["amount", "so_tien", "value", "tien"],
        "id_field_hints": ["transaction_id", "txn_id", "id", "ref_id", "reference"],
        "datetime_field_hints": ["transaction_date", "txn_date", "date", "time", "created_at"],
    },
    "Alipay": {
        "file_identifier": "a111275800000002",
        "cutoff_time": time(23, 0, 0),
        "buffer_window_minutes": 0,
        "file_delivery_type": "CONSOLIDATED",
        "tolerance_vnd": 0,
        "cancel_key_field": "merchant_txn_id",
        "amount_field_hints": ["amount", "so_tien", "value", "tien"],
        "id_field_hints": ["transaction_id", "txn_id", "id", "ref_id", "reference"],
        "datetime_field_hints": ["transaction_date", "txn_date", "date", "time", "created_at"],
    },
    "Tenpay": {
        "file_identifier": "tgp",
        "cutoff_time": time(22, 50, 0),
        "buffer_window_minutes": 1,
        "file_delivery_type": "CONSOLIDATED",
        "tolerance_vnd": 0,
        "cancel_key_field": "merchant_txn_id",
        "amount_field_hints": ["amount", "so_tien", "value", "tien"],
        "id_field_hints": ["transaction_id", "txn_id", "id", "ref_id", "reference"],
        "datetime_field_hints": ["transaction_date", "txn_date", "date", "time", "created_at"],
    },
}

DATA_DIR = Path("data")
PENDING_DIR = DATA_DIR / "pending_pool"
OUTPUT_DIR = DATA_DIR / "output"

# Excel color palette
COLOR_GREEN = "C6EFCE"
COLOR_RED = "FFC7CE"
COLOR_YELLOW = "FFEB9C"
COLOR_BLUE = "BDD7EE"
COLOR_ORANGE = "FCE4D6"
COLOR_HEADER = "1F4E79"


# ─────────────────────────────────────────────────────────────────────────────
# UTILITY HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def normalize_col_name(col: str) -> str:
    """Lowercase, strip, replace spaces/dashes with underscore."""
    return col.strip().lower().replace(" ", "_").replace("-", "_").replace(".", "_")


def detect_best_column(df: pd.DataFrame, hints: List[str]) -> Optional[str]:
    """Find the column in df that best matches a list of hints."""
    normalized = {normalize_col_name(c): c for c in df.columns}
    for hint in hints:
        if hint in normalized:
            return normalized[hint]
    # Partial match
    for hint in hints:
        for norm_col, orig_col in normalized.items():
            if hint in norm_col:
                return orig_col
    return None


def normalize_amount(series: pd.Series) -> pd.Series:
    """Clean amount column: remove commas, spaces, convert to float."""
    if series.dtype == object:
        cleaned = (
            series.astype(str)
            .str.replace(",", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.strip()
        )
        return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)
    return series.fillna(0.0)


def normalize_id(series: pd.Series) -> pd.Series:
    """Strip whitespace from ID column."""
    return series.astype(str).str.strip()


def read_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Read CSV, XLSX, or TXT file into DataFrame with auto encoding detection."""
    ext = Path(filename).suffix.lower()

    if ext == ".xlsx" or ext == ".xls":
        return pd.read_excel(io.BytesIO(file_bytes), dtype=str)

    # Try different encodings for CSV/TXT
    for encoding in ["utf-8-sig", "utf-8", "cp1252", "latin-1"]:
        try:
            content = file_bytes.decode(encoding)
            break
        except (UnicodeDecodeError, Exception):
            continue
    else:
        content = file_bytes.decode("utf-8", errors="replace")

    buf = io.StringIO(content)

    if ext == ".csv":
        return pd.read_csv(buf, dtype=str)
    elif ext == ".txt":
        # Try tab then pipe separator
        for sep in ["\t", "|", ";"]:
            try:
                buf.seek(0)
                df = pd.read_csv(buf, sep=sep, dtype=str)
                if len(df.columns) > 1:
                    return df
            except Exception:
                pass
        buf.seek(0)
        return pd.read_csv(buf, dtype=str)
    else:
        return pd.read_csv(buf, dtype=str)


# ─────────────────────────────────────────────────────────────────────────────
# PENDING POOL
# ─────────────────────────────────────────────────────────────────────────────

def load_pending_pool(partner: str) -> pd.DataFrame:
    """Load pending pool from JSON file."""
    PENDING_DIR.mkdir(parents=True, exist_ok=True)
    pool_file = PENDING_DIR / f"{partner}.json"
    if not pool_file.exists():
        return pd.DataFrame()
    try:
        with open(pool_file, "r", encoding="utf-8") as f:
            records = json.load(f)
        if not records:
            return pd.DataFrame()
        df = pd.DataFrame(records)
        return df
    except Exception:
        return pd.DataFrame()


def save_pending_pool(partner: str, df: pd.DataFrame):
    """Save pending pool to JSON file."""
    PENDING_DIR.mkdir(parents=True, exist_ok=True)
    pool_file = PENDING_DIR / f"{partner}.json"
    if df.empty:
        records = []
    else:
        records = df.to_dict(orient="records")
    with open(pool_file, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, default=str, indent=2)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN RECONCILIATION ENGINE
# ─────────────────────────────────────────────────────────────────────────────

class ReconciliationEngine:

    def __init__(self):
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        PENDING_DIR.mkdir(parents=True, exist_ok=True)

    # ── Step 1: Partner detection ────────────────────────────────────────────

    def detect_partner(self, filenames: List[str]) -> Optional[str]:
        """Detect partner name from uploaded filenames."""
        combined = " ".join(filenames).lower()
        for partner, cfg in PARTNER_CONFIG.items():
            if cfg["file_identifier"] in combined:
                return partner
        return None

    def separate_files(
        self,
        files: List[Tuple[str, bytes]],
        partner: str,
    ) -> Tuple[List[Tuple[str, bytes]], List[Tuple[str, bytes]]]:
        """Split uploaded files into external (partner) and internal."""
        identifier = PARTNER_CONFIG[partner]["file_identifier"]
        external, internal = [], []
        for fname, fbytes in files:
            if identifier in fname.lower():
                external.append((fname, fbytes))
            else:
                internal.append((fname, fbytes))
        return external, internal

    # ── Step 0: Consolidate DAILY_SPLIT files ────────────────────────────────

    def consolidate_files(
        self, file_list: List[Tuple[str, bytes]]
    ) -> Tuple[pd.DataFrame, List[str]]:
        """
        Merge multiple files into one DataFrame.
        Returns (merged_df, list_of_warnings).
        """
        if not file_list:
            return pd.DataFrame(), []

        dfs = []
        warnings = []
        reference_cols = None

        for fname, fbytes in file_list:
            try:
                df = read_file(fbytes, fname)
                df.columns = [c.strip() for c in df.columns]
                # Remove fully empty rows
                df = df.dropna(how="all")

                if reference_cols is None:
                    reference_cols = set(df.columns)
                else:
                    if set(df.columns) != reference_cols:
                        warnings.append(
                            f"⚠️ File '{fname}' có cấu trúc cột khác — "
                            f"kỳ vọng {len(reference_cols)} cột, thực tế {len(df.columns)} cột."
                        )
                        return pd.DataFrame(), warnings

                df["_source_file"] = fname
                dfs.append(df)
            except Exception as e:
                warnings.append(f"⚠️ Không đọc được file '{fname}': {str(e)}")

        if not dfs:
            return pd.DataFrame(), warnings

        merged = pd.concat(dfs, ignore_index=True)
        return merged, warnings

    # ── Gap Detection for DAILY_SPLIT ─────────────────────────────────────────

    def check_daily_split_gap(
        self,
        external_df: pd.DataFrame,
        partner: str,
        datetime_col: str,
    ) -> List[str]:
        """RULE 6: Check for gaps in date sequence for DAILY_SPLIT partners."""
        cfg = PARTNER_CONFIG[partner]
        if cfg["file_delivery_type"] != "DAILY_SPLIT":
            return []

        if datetime_col not in external_df.columns:
            return []

        try:
            dates = pd.to_datetime(
                external_df[datetime_col], errors="coerce"
            ).dropna().dt.date.unique()
            if len(dates) < 2:
                return []
            dates_sorted = sorted(dates)
            gaps = []
            for i in range(1, len(dates_sorted)):
                expected = dates_sorted[i - 1] + timedelta(days=1)
                actual = dates_sorted[i]
                # Skip weekends for gap check (optional — currently strict)
                if actual != expected:
                    gap_dates = []
                    d = expected
                    while d < actual:
                        gap_dates.append(str(d))
                        d += timedelta(days=1)
                    if gap_dates:
                        gaps.append(
                            f"🚫 RULE 6 — Thiếu file ngày: {', '.join(gap_dates)}. "
                            "Vui lòng bổ sung trước khi tiếp tục."
                        )
            return gaps
        except Exception:
            return []

    # ── Step 2: Data normalization ────────────────────────────────────────────

    def normalize_dataframe(
        self, df: pd.DataFrame, partner: str, is_internal: bool = False
    ) -> Tuple[pd.DataFrame, str, str, str]:
        """
        Normalize a DataFrame.
        Returns (normalized_df, id_col, amount_col, datetime_col).
        """
        cfg = PARTNER_CONFIG[partner]
        df = df.copy()

        # Detect key columns
        id_col = detect_best_column(df, cfg["id_field_hints"])
        amount_col = detect_best_column(df, cfg["amount_field_hints"])
        datetime_col = detect_best_column(df, cfg["datetime_field_hints"])

        # Fallback column names if not found
        if id_col is None:
            id_col = df.columns[0]
        if amount_col is None:
            # Pick first numeric-looking column
            for col in df.columns:
                try:
                    pd.to_numeric(df[col].dropna().head(5), errors="raise")
                    amount_col = col
                    break
                except Exception:
                    pass
            if amount_col is None:
                amount_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        if datetime_col is None and len(df.columns) > 2:
            datetime_col = df.columns[2]

        # Normalize ID
        df[id_col] = normalize_id(df[id_col])

        # Normalize Amount
        df[amount_col] = normalize_amount(df[amount_col])

        # Normalize datetime
        if datetime_col and datetime_col in df.columns:
            df[datetime_col] = pd.to_datetime(df[datetime_col], errors="coerce")

        # Add metadata columns
        df["_is_internal"] = is_internal
        df["_matched"] = False

        return df, id_col, amount_col, datetime_col

    # ── Step 2.5: Cancellation filtering ─────────────────────────────────────

    def filter_cancellations(
        self,
        df: pd.DataFrame,
        partner: str,
        amount_col: str,
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Separate cancelled transactions (pairs that sum to 0).
        Returns (clean_df, cancelled_df).
        """
        cfg = PARTNER_CONFIG[partner]
        cancel_key_raw = cfg["cancel_key_field"]

        # Find the actual column name (case-insensitive)
        cancel_key = None
        for col in df.columns:
            if normalize_col_name(col) == cancel_key_raw:
                cancel_key = col
                break

        if cancel_key is None:
            # Try partial match
            for col in df.columns:
                if "merchant" in normalize_col_name(col) or "txn" in normalize_col_name(col):
                    cancel_key = col
                    break

        if cancel_key is None:
            # No cancel key found — return as-is
            return df, pd.DataFrame(columns=df.columns)

        # Group by cancel key, compute sum of amounts
        df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0)
        grouped = df.groupby(cancel_key)[amount_col].agg(["sum", "count"])
        cancelled_keys = grouped[
            (grouped["count"] >= 2) & (grouped["sum"].abs() < 0.01)
        ].index

        cancelled_mask = df[cancel_key].isin(cancelled_keys)
        cancelled_df = df[cancelled_mask].copy()
        clean_df = df[~cancelled_mask].copy()

        return clean_df, cancelled_df

    # ── Step 3: Waterfall matching ─────────────────────────────────────────────

    def waterfall_match(
        self,
        internal_df: pd.DataFrame,
        external_df: pd.DataFrame,
        pending_df: pd.DataFrame,
        partner: str,
        internal_id_col: str,
        external_id_col: str,
        internal_amount_col: str,
        external_amount_col: str,
        internal_datetime_col: Optional[str],
    ) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Two-pass waterfall matching.
        Returns (matched_df, pending_to_add_df, discrepancy_df, external_only_df)
        """
        cfg = PARTNER_CONFIG[partner]
        tolerance = cfg["tolerance_vnd"]
        cutoff = cfg["cutoff_time"]
        buffer_minutes = cfg["buffer_window_minutes"]

        # Combine internal with pending pool
        if not pending_df.empty and internal_id_col in pending_df.columns:
            # Align columns
            all_cols = list(internal_df.columns)
            for c in pending_df.columns:
                if c not in all_cols:
                    all_cols.append(c)
            internal_combined = pd.concat(
                [internal_df, pending_df.reindex(columns=all_cols)],
                ignore_index=True,
            )
        else:
            internal_combined = internal_df.copy()

        # Build external lookup
        ext_lookup = {}
        for _, row in external_df.iterrows():
            key = str(row[external_id_col]).strip()
            ext_lookup.setdefault(key, []).append(row)

        matched_rows = []
        pending_new_rows = []
        discrepancy_rows = []
        used_external_ids = set()

        for _, int_row in internal_combined.iterrows():
            uid = str(int_row[internal_id_col]).strip()

            # Lượt 1: Exact match on ID
            if uid in ext_lookup:
                ext_candidates = ext_lookup[uid]
                for ext_row in ext_candidates:
                    if uid in used_external_ids:
                        continue
                    ext_amt = float(ext_row[external_amount_col]) if external_amount_col in ext_row.index else 0
                    int_amt = float(int_row[internal_amount_col]) if pd.notna(int_row.get(internal_amount_col, 0)) else 0
                    diff = abs(int_amt - ext_amt)

                    if diff <= tolerance:
                        # Full match
                        merged = {}
                        for c in internal_df.columns:
                            merged[f"int_{c}"] = int_row.get(c)
                        for c in external_df.columns:
                            merged[f"ext_{c}"] = ext_row.get(c)
                        merged["_match_type"] = "EXACT"
                        merged["_amount_diff"] = diff
                        matched_rows.append(merged)
                        used_external_ids.add(uid)
                        break
                    elif diff <= max(tolerance * 2, 1000):
                        # Tolerance match (fee difference)
                        merged = {}
                        for c in internal_df.columns:
                            merged[f"int_{c}"] = int_row.get(c)
                        for c in external_df.columns:
                            merged[f"ext_{c}"] = ext_row.get(c)
                        merged["_match_type"] = "TOLERANCE"
                        merged["_amount_diff"] = diff
                        matched_rows.append(merged)
                        used_external_ids.add(uid)
                        break
                    else:
                        # ID matched but amount too different → discrepancy
                        disc = dict(int_row)
                        disc["_reason"] = f"Amount lệch {diff:,.0f} VND"
                        disc["_ext_amount"] = ext_amt
                        disc["_int_amount"] = int_amt
                        disc["_match_type"] = "AMOUNT_MISMATCH"
                        disc["Ngay_GD_Goc"] = (
                            int_row.get(internal_datetime_col) if internal_datetime_col else ""
                        )
                        discrepancy_rows.append(disc)
                        used_external_ids.add(uid)
                        break
            else:
                # Not found in external — classify by buffer window
                txn_time = None
                if internal_datetime_col and internal_datetime_col in internal_combined.columns:
                    raw_time = int_row.get(internal_datetime_col)
                    if pd.notna(raw_time):
                        try:
                            txn_time = pd.to_datetime(raw_time).time()
                        except Exception:
                            txn_time = None

                classification = self._classify_by_buffer(
                    txn_time, cutoff, buffer_minutes, cfg["file_delivery_type"]
                )

                if classification == "MISSING_EXTERNAL":
                    disc = dict(int_row)
                    disc["_reason"] = "Không có trên file Đối tác (trước giờ Cut-off)"
                    disc["_match_type"] = "MISSING_EXTERNAL"
                    disc["Ngay_GD_Goc"] = (
                        int_row.get(internal_datetime_col) if internal_datetime_col else ""
                    )
                    discrepancy_rows.append(disc)
                else:
                    # PENDING_POOL
                    pend = dict(int_row)
                    pend["_partner"] = partner
                    pend["_pending_since"] = str(datetime.now().date())
                    pend["_reason"] = classification
                    pending_new_rows.append(pend)

        # External-only (on external but not in internal+pending)
        external_only_rows = []
        for _, ext_row in external_df.iterrows():
            uid = str(ext_row[external_id_col]).strip()
            if uid not in used_external_ids:
                row_dict = dict(ext_row)
                row_dict["_reason"] = "Có trên file Đối tác, không có trong Nội bộ"
                row_dict["_match_type"] = "EXTERNAL_ONLY"
                external_only_rows.append(row_dict)

        matched_df = pd.DataFrame(matched_rows) if matched_rows else pd.DataFrame()
        pending_new_df = pd.DataFrame(pending_new_rows) if pending_new_rows else pd.DataFrame()
        discrepancy_df = pd.DataFrame(discrepancy_rows) if discrepancy_rows else pd.DataFrame()
        external_only_df = pd.DataFrame(external_only_rows) if external_only_rows else pd.DataFrame()

        return matched_df, pending_new_df, discrepancy_df, external_only_df

    def _classify_by_buffer(
        self,
        txn_time: Optional[time],
        cutoff: Optional[time],
        buffer_minutes: int,
        delivery_type: str,
    ) -> str:
        """Classify an unmatched internal transaction."""
        if cutoff is None:
            # Apple: file-based, any unmatched = MISSING_EXTERNAL
            return "MISSING_EXTERNAL"

        if txn_time is None:
            # No time info → safe to pending
            return f"Lệch ca (không có thời gian giao dịch) → Pending Pool"

        buffer_dt = (
            datetime.combine(datetime.today(), cutoff) + timedelta(minutes=buffer_minutes)
        ).time()

        if txn_time < cutoff:
            return "MISSING_EXTERNAL"
        elif txn_time <= buffer_dt:
            return f"Vùng Buffer ({cutoff}–{buffer_dt}) → Pending Pool"
        else:
            return f"Lệch ca chốt sổ (sau {buffer_dt}) → Pending Pool"

    # ── Step 4: Update Pending Pool ───────────────────────────────────────────

    def update_pending_pool(
        self,
        partner: str,
        existing_pool: pd.DataFrame,
        matched_df: pd.DataFrame,
        new_pending_df: pd.DataFrame,
        id_col: str,
    ) -> pd.DataFrame:
        """
        Remove matched rows from pool, add new pending rows.
        Returns updated pool DataFrame.
        """
        if not existing_pool.empty and not matched_df.empty:
            matched_ids = set()
            for col in [f"int_{id_col}", id_col]:
                if col in matched_df.columns:
                    matched_ids.update(matched_df[col].astype(str).tolist())
            if id_col in existing_pool.columns:
                existing_pool = existing_pool[
                    ~existing_pool[id_col].astype(str).isin(matched_ids)
                ]

        if not new_pending_df.empty:
            if existing_pool.empty:
                updated = new_pending_df
            else:
                updated = pd.concat([existing_pool, new_pending_df], ignore_index=True)
        else:
            updated = existing_pool

        return updated

    # ── Excel Export ──────────────────────────────────────────────────────────

    def generate_excel(
        self,
        partner: str,
        matched_df: pd.DataFrame,
        cancelled_df: pd.DataFrame,
        pending_df: pd.DataFrame,
        discrepancy_df: pd.DataFrame,
        summary: dict,
    ) -> str:
        """Generate Excel output with 4 tabs. Returns file path."""
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Ket_Qua_Doi_Soat_{partner}_{timestamp}.xlsx"
        filepath = OUTPUT_DIR / filename

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Tab 1: Matched
        self._write_tab(
            wb,
            "Tab_Khớp",
            matched_df,
            header_color=COLOR_GREEN,
            tab_color="00B050",
        )

        # Tab 2: Cancelled
        self._write_tab(
            wb,
            "Tab_Đã_Hủy_Bỏ_Qua",
            cancelled_df,
            header_color=COLOR_BLUE,
            tab_color="0070C0",
        )

        # Tab 3: Pending
        self._write_tab(
            wb,
            "Tab_Vùng_Chờ_Pending",
            pending_df,
            header_color=COLOR_YELLOW,
            tab_color="FFC000",
        )

        # Tab 4: Discrepancy (most important)
        self._write_tab(
            wb,
            "Tab_Lệch_Nghi_Vấn",
            discrepancy_df,
            header_color=COLOR_RED,
            tab_color="FF0000",
        )

        # Tab 5: Summary dashboard
        self._write_summary_tab(wb, partner, summary)

        wb.save(str(filepath))
        return str(filepath)

    def _write_tab(
        self,
        wb: openpyxl.Workbook,
        sheet_name: str,
        df: pd.DataFrame,
        header_color: str,
        tab_color: str,
    ):
        """Write a DataFrame to a styled worksheet."""
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_properties.tabColor = tab_color

        if df.empty:
            ws.append(["(Không có dữ liệu)"])
            return

        # Write headers
        header_fill = PatternFill("solid", fgColor=header_color)
        header_font = Font(bold=True, color="000000")
        headers = list(df.columns)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write data
        row_fill_even = PatternFill("solid", fgColor="F2F2F2")
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            fill = row_fill_even if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Clean value for Excel
                if pd.isna(value) if not isinstance(value, str) else False:
                    cell.value = ""
                elif isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.value = value
                else:
                    cell.value = str(value) if value is not None else ""
                if fill:
                    cell.fill = fill

        # Auto-fit column widths (approximate)
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col_cells:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # Freeze top row
        ws.freeze_panes = "A2"

    def _write_summary_tab(self, wb: openpyxl.Workbook, partner: str, summary: dict):
        """Write a summary dashboard tab."""
        ws = wb.create_sheet(title="📊 Tóm Tắt", index=0)
        ws.sheet_properties.tabColor = "1F4E79"

        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        header_font = Font(bold=True, color="FFFFFF", size=12)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

        # Title
        ws.merge_cells("A1:B1")
        title_cell = ws.cell(row=1, column=1, value=f"KẾT QUẢ ĐỐI SOÁT — {partner}")
        title_cell.fill = header_fill
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal="center")

        rows = [
            ("Thời gian xuất báo cáo", summary.get("report_time", "")),
            ("Đối tác", partner),
            ("Kỳ đối soát", summary.get("period", "")),
            ("", ""),
            ("Tổng dòng Nội bộ", summary.get("total_internal", 0)),
            ("Tổng dòng Đối tác", summary.get("total_external", 0)),
            ("", ""),
            ("✅ Khớp hoàn toàn", summary.get("matched", 0)),
            ("🚫 Tự động loại trừ (Cancel)", summary.get("cancelled", 0)),
            ("⏳ Treo Pending Pool", summary.get("pending", 0)),
            ("⚠️  Lệch Nghi Vấn (cần xử lý)", summary.get("discrepancy", 0)),
        ]

        fill_white = PatternFill("solid", fgColor="FFFFFF")
        fill_alt = PatternFill("solid", fgColor="EBF3FB")

        for i, (label, value) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=label).font = Font(bold=bool(label))
            ws.cell(row=i, column=2, value=value)
            fill = fill_alt if i % 2 == 0 else fill_white
            for col in [1, 2]:
                ws.cell(row=i, column=col).fill = fill

        # Highlight discrepancy row in red if > 0
        if summary.get("discrepancy", 0) > 0:
            disc_row = 12  # row index for discrepancy
            for col in [1, 2]:
                ws.cell(row=disc_row, column=col).fill = PatternFill("solid", fgColor=COLOR_RED)
                ws.cell(row=disc_row, column=col).font = Font(bold=True)

    # ── Main Process Orchestrator ─────────────────────────────────────────────

    def process(
        self,
        files: List[Tuple[str, bytes]],
        partner: Optional[str] = None,
    ) -> dict:
        """
        Full reconciliation pipeline.
        Returns a result dict with summary + file path.
        """
        result = {
            "success": False,
            "partner": None,
            "warnings": [],
            "errors": [],
            "summary": {},
            "excel_path": None,
        }

        # Step 1: Partner detection
        filenames = [f[0] for f in files]
        if partner is None:
            partner = self.detect_partner(filenames)
        if partner is None:
            result["errors"].append(
                "Không nhận diện được đối tác từ tên file. "
                "Vui lòng cho biết đây là file của đối tác nào (Apple / Alipay / Tenpay)."
            )
            return result

        result["partner"] = partner
        cfg = PARTNER_CONFIG[partner]

        # Separate external vs internal files
        external_files, internal_files = self.separate_files(files, partner)

        if not external_files:
            result["errors"].append(
                f"Không tìm thấy file của đối tác {partner} "
                f"(cần tên file chứa '{cfg['file_identifier']}')."
            )
            return result

        if not internal_files:
            result["errors"].append(
                "Không tìm thấy file Nội bộ. "
                "Vui lòng upload cả file xuất từ hệ thống nội bộ."
            )
            return result

        # Step 0: Consolidate external files
        external_df, ext_warnings = self.consolidate_files(external_files)
        result["warnings"].extend(ext_warnings)
        if external_df.empty and ext_warnings:
            result["errors"].extend(ext_warnings)
            return result

        # Step 0: Consolidate internal files
        internal_df, int_warnings = self.consolidate_files(internal_files)
        result["warnings"].extend(int_warnings)
        if internal_df.empty and int_warnings:
            result["errors"].extend(int_warnings)
            return result

        # Step 2: Normalize
        ext_df, ext_id_col, ext_amt_col, ext_dt_col = self.normalize_dataframe(
            external_df, partner, is_internal=False
        )
        int_df, int_id_col, int_amt_col, int_dt_col = self.normalize_dataframe(
            internal_df, partner, is_internal=True
        )

        # RULE 6: Gap detection for DAILY_SPLIT
        if ext_dt_col:
            gaps = self.check_daily_split_gap(ext_df, partner, ext_dt_col)
            if gaps:
                result["errors"].extend(gaps)
                return result

        # Get period
        period_str = ""
        if ext_dt_col and ext_dt_col in ext_df.columns:
            valid_dates = pd.to_datetime(ext_df[ext_dt_col], errors="coerce").dropna()
            if not valid_dates.empty:
                period_str = (
                    f"{valid_dates.min().strftime('%d/%m/%Y')} → "
                    f"{valid_dates.max().strftime('%d/%m/%Y')}"
                )

        # Step 2.5: Filter cancellations from BOTH datasets
        ext_clean, ext_cancelled = self.filter_cancellations(ext_df, partner, ext_amt_col)
        int_clean, int_cancelled = self.filter_cancellations(int_df, partner, int_amt_col)
        cancelled_combined = pd.concat([ext_cancelled, int_cancelled], ignore_index=True)

        # Load pending pool
        pending_pool = load_pending_pool(partner)
        pending_count_before = len(pending_pool) if not pending_pool.empty else 0

        # Step 3: Waterfall matching
        matched_df, new_pending_df, discrepancy_df, ext_only_df = self.waterfall_match(
            int_clean, ext_clean, pending_pool, partner,
            int_id_col, ext_id_col, int_amt_col, ext_amt_col, int_dt_col
        )

        # Add external-only to discrepancy
        if not ext_only_df.empty:
            discrepancy_df = pd.concat([discrepancy_df, ext_only_df], ignore_index=True)

        # Step 4: Update pending pool
        updated_pool = self.update_pending_pool(
            partner, pending_pool, matched_df, new_pending_df, int_id_col
        )
        save_pending_pool(partner, updated_pool)

        # Build summary
        summary = {
            "report_time": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "period": period_str,
            "total_internal": len(int_df),
            "total_external": len(ext_df),
            "matched": len(matched_df),
            "cancelled": len(cancelled_combined),
            "pending": len(new_pending_df),
            "discrepancy": len(discrepancy_df),
            "pending_pool_before": pending_count_before,
            "pending_pool_after": len(updated_pool) if not updated_pool.empty else 0,
        }

        # Generate Excel
        excel_path = self.generate_excel(
            partner, matched_df, cancelled_combined,
            new_pending_df, discrepancy_df, summary
        )

        result.update({
            "success": True,
            "summary": summary,
            "excel_path": excel_path,
            "excel_filename": Path(excel_path).name,
        })
        return result
